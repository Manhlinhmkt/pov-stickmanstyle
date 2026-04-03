#!/usr/bin/env python3
"""
Export Episode Assets
=====================
Exports VO script to .txt and image prompts to .xlsx

Usage:
    python scripts/export_episode_assets.py episodes/PP_0001
    python scripts/export_episode_assets.py PP_0001
"""

import csv
import glob
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Error: openpyxl không được cài đặt.")
    print("Chạy: pip install openpyxl")
    sys.exit(1)


def fix_csv_encoding(episode_dir: Path) -> int:
    """Re-encode all CSV files in episode dir to UTF-8 BOM for Excel compatibility."""
    import os
    csvs = glob.glob(str(episode_dir / "*.csv"))
    count = 0
    for f in csvs:
        try:
            with open(f, "r", encoding="utf-8") as fh:
                content = fh.read()
            content = content.lstrip("\ufeff")
            # Try direct write first
            try:
                with open(f, "w", encoding="utf-8-sig", newline="") as fh:
                    fh.write(content)
            except PermissionError:
                # File locked — use temp file approach
                tmp = f + ".tmp"
                with open(tmp, "w", encoding="utf-8-sig", newline="") as fh:
                    fh.write(content)
                try:
                    os.replace(tmp, f)
                except PermissionError:
                    os.remove(tmp)
                    print(f"  ⚠️  Skipped (locked): {os.path.basename(f)}")
                    continue
            count += 1
        except Exception as e:
            print(f"  ⚠️  Error on {os.path.basename(f)}: {e}")
    return count


def export_vo_script(episode_dir: Path) -> Path:
    """Export VO_EN column to .txt file with double line breaks."""
    episode_id = episode_dir.name
    vo_script_path = episode_dir / "vo_script_table.csv"
    output_path = episode_dir / f"{episode_id}.txt"
    
    if not vo_script_path.exists():
        raise FileNotFoundError(f"Không tìm thấy: {vo_script_path}")
    
    lines = []
    with open(vo_script_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            vo_text = row.get('VO_EN', '').strip()
            if vo_text:
                lines.append(vo_text)
    
    # Write with double line breaks (extra Enter between each line)
    with open(output_path, 'w', encoding='utf-8-sig') as f:
        f.write('\n\n'.join(lines))
    
    return output_path





def export_image_prompts(episode_dir: Path) -> Path:
    """Export Image_Prompt column to .xlsx with STT, Prompt, and Images columns."""
    import re
    import os
    episode_id = episode_dir.name
    prompts_path = episode_dir / "image_prompts.csv"
    output_path = episode_dir / f"{episode_id}.xlsx"
    
    REF_IMAGES_BASE = r"D:\Veo3\Image\ref_images"
    EPISODE_IMAGES_BASE = r"A:\Veo3\Image"
    DEFAULT_IMAGE = rf"{REF_IMAGES_BASE}\white.PNG"
    
    if not prompts_path.exists():
        raise FileNotFoundError(f"Không tìm thấy: {prompts_path}")
    
    # Read prompts and ref images
    rows = []
    with open(prompts_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            prompt = row.get('Image_Prompt', '').strip()
            ref_images = row.get('Ref_Images', '').strip()
            veo3_prompt = row.get('Veo3_Prompt', '').strip()
            if prompt:
                rows.append((prompt, ref_images, veo3_prompt))
    
    def resolve_ref_path(ref_name: str, ep_id: str) -> str:
        """Resolve a single ref image name to full path."""
        ref_name = ref_name.strip()
        if not ref_name:
            return DEFAULT_IMAGE
        # Sub-frame refs: numeric pattern like "2_1.png", "276_1.png"
        if re.match(r'^\d+_\d+\.png$', ref_name):
            return rf"{EPISODE_IMAGES_BASE}\{ep_id}\{ref_name}"
        # Named ref images: human_girl.png, pet_dog_golden.png, etc.
        return rf"{REF_IMAGES_BASE}\{ref_name}"
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Prompts"
    
    # Header styling
    header_font = Font(bold=True)
    ws['A1'] = 'STT'
    ws['B1'] = 'Prompt'
    ws['C1'] = 'Images'
    ws['D1'] = 'Generated'
    ws['E1'] = 'Veo3 Prompt'
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font
    ws['E1'].font = header_font
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 120
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 100
    
    # Data rows
    generated_count = 0
    veo3_count = 0
    for idx, (prompt, ref_images, veo3_prompt) in enumerate(rows, start=1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=prompt)
        
        # Resolve ref images to full paths
        if ref_images:
            refs = ref_images.split('|')
            full_paths = [resolve_ref_path(r, episode_id) for r in refs]
            images_value = ';'.join(full_paths)
        else:
            images_value = DEFAULT_IMAGE
        
        ws.cell(row=idx + 1, column=3, value=images_value)
        
        # Check if generated image exists
        generated_path = os.path.join(EPISODE_IMAGES_BASE, episode_id, f"{idx}_1.png")
        if os.path.exists(generated_path):
            ws.cell(row=idx + 1, column=4, value='x')
            generated_count += 1
        
        # Veo 3 animation prompt
        if veo3_prompt:
            ws.cell(row=idx + 1, column=5, value=veo3_prompt)
            veo3_count += 1
    
    wb.save(output_path)
    print(f"   📊 {generated_count}/{len(rows)} images already generated")
    if veo3_count > 0:
        print(f"   🎬 {veo3_count} Veo 3 animations suggested")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/export_episode_assets.py <episode_path>")
        print("Example: python scripts/export_episode_assets.py episodes/PP_0001")
        sys.exit(1)
    
    # Parse episode path
    input_path = sys.argv[1]
    
    # Handle different input formats
    if input_path.startswith("episodes/") or input_path.startswith("episodes\\"):
        episode_dir = Path(input_path)
    elif input_path.startswith("PP_") or input_path.startswith("EC_"):
        episode_dir = Path("episodes") / input_path
    else:
        episode_dir = Path(input_path)
    
    if not episode_dir.exists():
        print(f"Error: Không tìm thấy folder: {episode_dir}")
        sys.exit(1)
    
    episode_id = episode_dir.name
    print(f"\n{'='*60}")
    print(f"EXPORT EPISODE ASSETS: {episode_id}")
    print(f"{'='*60}")
    
    # Fix CSV encoding (UTF-8 BOM) before any export
    fixed = fix_csv_encoding(episode_dir)
    print(f"✅ Encoding fix → {fixed} CSV files re-encoded (UTF-8 BOM)")
    
    # Export VO script (English)
    try:
        txt_path = export_vo_script(episode_dir)
        print(f"✅ VO Script (EN)  → {txt_path}")
    except FileNotFoundError as e:
        print(f"⚠️  VO Script (EN)  → {e}")
    
    
    # Export image prompts
    try:
        xlsx_path = export_image_prompts(episode_dir)
        print(f"✅ Prompts    → {xlsx_path}")
    except FileNotFoundError as e:
        print(f"⚠️  Prompts    → {e}")
    
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
